import os
import pyexcel as p
import openpyxl, datetime


cwd = os.getcwd()
for file in os.listdir(cwd):
    if file.endswith(".xls"):
        #newname = file.replace('xls', 'xlsx')
        filename = file.split('.')[0]
        p.save_book_as(file_name =file, dest_file_name=filename+'.xlsx')

write_wb = openpyxl.Workbook()
master_rows = write_wb.active
for file in os.listdir(cwd):
    if file.endswith('.xlsx'):
        filename = file.split('.')[0]
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        for row in ws:
            row_values = []
            for item in row:
                row_values.append(item.value)
            #row_values = [item.value for item in row]
            if type(row[0].value)==str:
                if row[0].value.startswith('Home Room'):
                    pass
                elif row[0].value=='Student' and row[-1].value=='Amount':
                    pass
                else:
                    master_rows.append(tuple(row_values+[filename]))
            elif type(row[0].value)==datetime.datetime:
                pass
            else:
                master_rows.append(tuple(row_values + [filename]))
write_wb.save('master.xlsx')

wb = openpyxl.load_workbook('master.xlsx')
ws = wb.active
latest_name = 0
for row in ws:
    #row_values = [item.value for item in row]
    if type(row[0].value)==str:
        latest_name+=1
        row[0].value = latest_name
    else:
        row[0].value = latest_name
wb.save('master_2.xlsx')